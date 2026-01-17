from django.shortcuts import render, redirect
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Q, Count
from django.utils import timezone
from datetime import datetime
from django.db import IntegrityError
from django.core.paginator import Paginator
from django.http import HttpResponse, JsonResponse
from django.conf import settings
from openpyxl import Workbook, load_workbook
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib import colors
from reportlab.lib.units import inch
from .models import Member

def login_view(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            return redirect('member_list')
        else:
            return render(request, 'members/login.html', {'error': 'Invalid credentials'})
    return render(request, 'members/login.html')

@login_required
def member_list(request):
    # Update active status for expired memberships
    from django.utils import timezone
    today = timezone.now().date()
    Member.objects.filter(membership_end_date__lt=today, is_active=True).update(is_active=False)

    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'add':
            name = request.POST.get('name')
            email = request.POST.get('email')
            phone_number = request.POST.get('phone_number')
            dob_str = request.POST.get('dob')
            gender = request.POST.get('gender')
            join_date_str = request.POST.get('join_date')
            membership_type = request.POST.get('membership_type')
            membership_start_date = request.POST.get('membership_start_date')
            membership_end_date = request.POST.get('membership_end_date')
            is_active = request.POST.get('is_active') == 'on'
            if not email.endswith('@gmail.com'):
                messages.error(request, 'Please enter valid email address.')
                return redirect('member_list')
            try:
                join_date = datetime.strptime(join_date_str, '%Y-%m-%d').date()
                if join_date > timezone.now().date():
                    messages.error(request, 'Join date cannot be in the future.')
                    return redirect('member_list')
            except ValueError:
                messages.error(request, 'Invalid join date format.')
                return redirect('member_list')
            dob = None
            if dob_str:
                try:
                    dob = datetime.strptime(dob_str, '%Y-%m-%d').date()
                except ValueError:
                    messages.error(request, 'Invalid DOB format.')
                    return redirect('member_list')
            try:
                Member.objects.create(
                    name=name, email=email, phone_number=phone_number, dob=dob, gender=gender, join_date=join_date, membership_type=membership_type,
                    membership_start_date=membership_start_date, membership_end_date=membership_end_date, is_active=is_active
                )
            except IntegrityError:
                messages.error(request, 'A member with this email already exists.')
                return redirect('member_list')
        elif action == 'update':
            member_id = request.POST.get('member_id')
            member = Member.objects.get(id=member_id)
            member.name = request.POST.get('name')
            email = request.POST.get('email')
            if not email.endswith('@gmail.com'):
                messages.error(request, 'Please enter valid email address.')
                return redirect('member_list')
            member.email = email
            member.phone_number = request.POST.get('phone_number')
            dob_str = request.POST.get('dob')
            dob = None
            if dob_str:
                try:
                    dob = datetime.strptime(dob_str, '%Y-%m-%d').date()
                except ValueError:
                    messages.error(request, 'Invalid DOB format.')
                    return redirect('member_list')
            member.dob = dob
            join_date_str = request.POST.get('join_date')
            try:
                join_date = datetime.strptime(join_date_str, '%Y-%m-%d').date()
                if join_date > timezone.now().date():
                    messages.error(request, 'Join date cannot be in the future.')
                    return redirect('member_list')
            except ValueError:
                messages.error(request, 'Invalid join date format.')
                return redirect('member_list')
            member.join_date = join_date
            member.membership_type = request.POST.get('membership_type')
            member.membership_start_date = request.POST.get('membership_start_date')
            member.membership_end_date = request.POST.get('membership_end_date')
            member.gender = request.POST.get('gender')
            member.is_active = request.POST.get('is_active') == 'on'
            member.save()
        elif action == 'delete':
            member_id = request.POST.get('member_id')
            Member.objects.filter(id=member_id).delete()
        return redirect('member_list')
    search_query = request.GET.get('search', '')
    sort_by = request.GET.get('sort_by', 'join_date')
    direction = request.GET.get('direction', 'desc')

    # Validate sort_by field
    allowed_sort_fields = ['id', 'name', 'email', 'join_date', 'membership_type', 'is_active']
    if sort_by not in allowed_sort_fields:
        sort_by = 'id'

    # Validate direction
    if direction not in ['asc', 'desc']:
        direction = 'asc'

    # Build order_by string
    order_by_field = f'-{sort_by}' if direction == 'desc' else sort_by

    if search_query:
        members = Member.objects.filter(
            Q(name__icontains=search_query) | Q(email__icontains=search_query)
        ).order_by(order_by_field)
    else:
        members = Member.objects.all().order_by(order_by_field)
    total_members = Member.objects.count()
    active_members = Member.objects.filter(is_active=True).count()
    inactive_members = total_members - active_members
    membership_types = Member.objects.values('membership_type').annotate(count=Count('membership_type')).order_by('-count')
    paginator = Paginator(members, 5)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    context = {
        'members': page_obj,
        'search_query': search_query,
        'sort_by': sort_by,
        'direction': direction,
        'total_members': total_members,
        'active_members': active_members,
        'inactive_members': inactive_members,
        'membership_types': membership_types,
    }
    return render(request, 'members/member_list.html', context)

def logout_view(request):
    logout(request)
    return redirect('login')

@login_required
def export_excel(request):
    members = Member.objects.all()
    wb = Workbook()
    ws = wb.active
    ws.title = "Members"

    # Add headers
    headers = ['Name', 'Email', 'Phone Number', 'Date of Birth', 'Gender', 'Join Date', 'Membership Type', 'Membership Start Date', 'Membership End Date', 'Is Active']
    for col_num, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_num, value=header)

    # Add data
    for row_num, member in enumerate(members, 2):
        ws.cell(row=row_num, column=1, value=member.name)
        ws.cell(row=row_num, column=2, value=member.email)
        ws.cell(row=row_num, column=3, value=member.phone_number)
        ws.cell(row=row_num, column=4, value=member.dob.strftime('%Y-%m-%d') if member.dob else '')
        ws.cell(row=row_num, column=5, value=member.gender)
        ws.cell(row=row_num, column=6, value=member.join_date.strftime('%Y-%m-%d'))
        ws.cell(row=row_num, column=7, value=member.membership_type)
        ws.cell(row=row_num, column=8, value=member.membership_start_date.strftime('%Y-%m-%d'))
        ws.cell(row=row_num, column=9, value=member.membership_end_date.strftime('%Y-%m-%d'))
        ws.cell(row=row_num, column=10, value='Yes' if member.is_active else 'No')

    # Create response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=members.xlsx'
    wb.save(response)
    return response

@login_required
def import_excel(request):
    if request.method == 'POST' and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']
        try:
            wb = load_workbook(excel_file)
            ws = wb.active
            imported_count = 0
            errors = []
            for row in ws.iter_rows(min_row=2, values_only=True):  # Skip header
                if len(row) < 10:
                    errors.append("Row has insufficient columns")
                    continue
                name, email, phone_number, dob_str, gender, join_date_str, membership_type, membership_start_date_str, membership_end_date_str, is_active_str = row[:10]
                if not name or not email:
                    errors.append(f"Name and email are required for row: {row}")
                    continue
                if not email.endswith('@gmail.com'):
                    errors.append(f"Invalid email for {name}: {email}")
                    continue
                try:
                    join_date = datetime.strptime(join_date_str, '%Y-%m-%d').date()
                    if join_date > timezone.now().date():
                        errors.append(f"Join date cannot be in the future for {name}")
                        continue
                except (ValueError, TypeError):
                    errors.append(f"Invalid join date format for {name}: {join_date_str}")
                    continue
                dob = None
                if dob_str:
                    try:
                        dob = datetime.strptime(dob_str, '%Y-%m-%d').date()
                    except (ValueError, TypeError):
                        errors.append(f"Invalid DOB format for {name}: {dob_str}")
                        continue
                try:
                    membership_start_date = datetime.strptime(membership_start_date_str, '%Y-%m-%d').date()
                except (ValueError, TypeError):
                    errors.append(f"Invalid membership start date format for {name}: {membership_start_date_str}")
                    continue
                try:
                    membership_end_date = datetime.strptime(membership_end_date_str, '%Y-%m-%d').date()
                except (ValueError, TypeError):
                    errors.append(f"Invalid membership end date format for {name}: {membership_end_date_str}")
                    continue
                is_active = is_active_str.lower() == 'yes' if is_active_str else False
                try:
                    Member.objects.create(
                        name=name, email=email, phone_number=phone_number or '', dob=dob, gender=gender or '',
                        join_date=join_date, membership_type=membership_type,
                        membership_start_date=membership_start_date, membership_end_date=membership_end_date,
                        is_active=is_active
                    )
                    imported_count += 1
                except IntegrityError:
                    errors.append(f"Member with email {email} already exists")
                    continue
                except Exception as e:
                    errors.append(f"Error creating member {name}: {str(e)}")
                    continue
            if imported_count > 0:
                messages.success(request, f"Successfully imported {imported_count} members.")
            if errors:
                for error in errors:
                    messages.error(request, error)
        except Exception as e:
            messages.error(request, f"Error processing file: {str(e)}")
    return redirect('member_list')

def add_member(request):
    if request.method == 'POST':
        name = request.POST.get('name')
        email = request.POST.get('email')
        phone_number = request.POST.get('phone_number')
        dob_str = request.POST.get('dob')
        gender = request.POST.get('gender')
        join_date_str = request.POST.get('join_date')
        membership_type = request.POST.get('membership_type')
        membership_start_date = request.POST.get('membership_start_date')
        membership_end_date = request.POST.get('membership_end_date')
        is_active = request.POST.get('is_active') == 'on'
        if not email.endswith('@gmail.com'):
            return render(request, 'members/add_member.html', {'error': 'Please enter valid email address.'})
        try:
            join_date = datetime.strptime(join_date_str, '%Y-%m-%d').date()
            if join_date > timezone.now().date():
                return render(request, 'members/add_member.html', {'error': 'Join date cannot be in the future.'})
        except ValueError:
            return render(request, 'members/add_member.html', {'error': 'Invalid join date format.'})
        dob = None
        if dob_str:
            try:
                dob = datetime.strptime(dob_str, '%Y-%m-%d').date()
            except ValueError:
                return render(request, 'members/add_member.html', {'error': 'Invalid DOB format.'})
        try:
            Member.objects.create(
                name=name, email=email, phone_number=phone_number, dob=dob, gender=gender, join_date=join_date, membership_type=membership_type,
                membership_start_date=membership_start_date, membership_end_date=membership_end_date, is_active=is_active
            )
        except IntegrityError:
            return render(request, 'members/add_member.html', {'error': 'A member with this email already exists.'})
        return redirect('member_list')
    return render(request, 'members/add_member.html')

@login_required
def generate_invoice_pdf(request, member_id):
    send_email = request.GET.get('send_email') == 'true' or request.POST.get('send_email') == 'on'
    try:
        member = Member.objects.get(id=member_id)
    except Member.DoesNotExist:
        messages.error(request, 'Member not found.')
        return redirect('member_list')

    # Define membership pricing
    membership_prices = {
        'strength': 50.00,
        'cardio': 40.00,
        'crossfit': 60.00,
    }

    # Check if POST data is provided (from modal)
    if request.method == 'POST':
        membership_start_date = request.POST.get('membership_start_date')
        membership_end_date = request.POST.get('membership_end_date')
        membership_type = request.POST.get('membership_type')
        is_active = request.POST.get('is_active') == 'on'
        payment_amount = request.POST.get('payment_amount')
        payment_mode = request.POST.get('payment_mode')
        transaction_id = request.POST.get('transaction_id')
        comments = request.POST.get('comments')

        # Parse dates
        try:
            membership_start_date = datetime.strptime(membership_start_date, '%Y-%m-%d').date()
            membership_end_date = datetime.strptime(membership_end_date, '%Y-%m-%d').date()
        except ValueError:
            messages.error(request, 'Invalid date format.')
            return redirect('member_list')

        # Update member data with the new values
        member.membership_start_date = membership_start_date
        member.membership_end_date = membership_end_date
        member.membership_type = membership_type
        member.is_active = is_active
        member.fees_amount = request.POST.get('fees_amount') or payment_amount
        member.payment_mode = request.POST.get('payment_mode')
        member.transaction_id = request.POST.get('transaction_id')
        member.comments = request.POST.get('comments')
        member.save()

        # Use POST data for calculations
        price_per_month = membership_prices.get(membership_type.lower(), 50.00)
        duration_days = (membership_end_date - membership_start_date).days
        duration_months = max(1, round(duration_days / 30))  # Approximate months
        total_amount = price_per_month * duration_months
    else:
        # Use member data (original behavior)
        membership_start_date = member.membership_start_date
        membership_end_date = member.membership_end_date
        membership_type = member.membership_type
        is_active = member.is_active
        payment_amount = None
        payment_mode = None
        transaction_id = None

        # Calculate charges
        price_per_month = membership_prices.get(member.membership_type.lower(), 50.00)
        duration_days = (member.membership_end_date - member.membership_start_date).days
        duration_months = max(1, round(duration_days / 30))  # Approximate months
        total_amount = price_per_month * duration_months

    # Generate PDF content
    from io import BytesIO

    pdf_buffer = BytesIO()
    doc = SimpleDocTemplate(pdf_buffer, pagesize=A4)
    styles = getSampleStyleSheet()

    # Custom styles
    title_style = ParagraphStyle(
        'Title',
        parent=styles['Heading1'],
        fontSize=24,
        spaceAfter=30,
        alignment=1  # Center
    )
    header_style = ParagraphStyle(
        'Header',
        parent=styles['Heading2'],
        fontSize=14,
        spaceAfter=20
    )
    normal_style = styles['Normal']

    # Build PDF content
    content = []

    # Company Header
    content.append(Paragraph("<b>A1 Lift and Fit hub unisex Gym</b>", ParagraphStyle('Company', parent=styles['Heading2'], fontSize=14, alignment=1, spaceAfter=10)))
    content.append(Paragraph("123 Fitness Street, Health City, India - 400001", ParagraphStyle('Address', parent=styles['Normal'], alignment=1, spaceAfter=5)))
    content.append(Paragraph("Phone: +91-9876543210 | Email: info@fitnessprogym.com", ParagraphStyle('Contact', parent=styles['Normal'], alignment=1, spaceAfter=10)))

    # Title
    # content.append(Paragraph("Gym Membership Invoice", title_style))
    # content.append(Spacer(1, 12))

    # Invoice details
    invoice_number = f"INV-{member.id}-{timezone.now().strftime('%Y%m%d')}"
    invoice_date = timezone.now().strftime('%B %d, %Y')

    content.append(Paragraph(f"<b>Invoice Number:</b> {invoice_number}", normal_style))
    content.append(Paragraph(f"<b>Invoice Date:</b> {invoice_date}", normal_style))
    content.append(Spacer(1, 10))

    # Member information
    content.append(Paragraph("Member Information", header_style))
    member_info = [
        ["Name:", member.name],
        ["Email:", member.email],
        ["Phone:", member.phone_number or 'N/A'],
        ["Date of Birth:", member.dob.strftime('%Y-%m-%d') if member.dob else 'N/A'],
        ["Gender:", member.gender or 'N/A'],
        ["Join Date:", member.join_date.strftime('%Y-%m-%d')],
    ]

    member_table = Table(member_info, colWidths=[2*inch, 4*inch])
    member_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    content.append(member_table)
    content.append(Spacer(1, 10))

    # Membership details
    content.append(Paragraph("Membership Details", header_style))
    membership_info = [
        ["Membership Type:", membership_type.title()],
        ["Start Date:", membership_start_date.strftime('%Y-%m-%d')],
        ["End Date:", membership_end_date.strftime('%Y-%m-%d')],
        ["Status:", "Active" if is_active else "Inactive"],
    ]

    membership_table = Table(membership_info, colWidths=[2*inch, 4*inch])
    membership_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    content.append(membership_table)
    content.append(Spacer(1, 12))

    # Payment Details (only if POST data is provided)
    if request.method == 'POST' and (member.fees_amount or member.payment_mode or member.transaction_id):
        content.append(Paragraph("Payment Details", header_style))
        payment_info = [
            ["Fees Amount:", f"Rupees {member.fees_amount} /-" if member.fees_amount else 'N/A'],
            ["Payment Mode:", member.payment_mode.title() if member.payment_mode else 'N/A'],
            ["Transaction ID:", member.transaction_id or 'N/A'],
        ]

        payment_table = Table(payment_info, colWidths=[2*inch, 4*inch])
        payment_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        content.append(payment_table)
        content.append(Spacer(1, 12))


    # Comments (only if provided)
    if request.method == 'POST' and comments:
        content.append(Paragraph("Comments", header_style))
        comments_info = [
            ["Comments:", comments],
        ]

        comments_table = Table(comments_info, colWidths=[2*inch, 4*inch])
        comments_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        content.append(comments_table)
        content.append(Spacer(1, 12))

    # Footer
    content.append(Paragraph("Terms & Conditions:", ParagraphStyle('FooterHeader', parent=styles['Heading3'], fontSize=12, spaceAfter=10)))
    # content.append(Paragraph("• Payment is due within 7 days of invoice date.", normal_style))
    # content.append(Paragraph("• Late payments may incur additional charges.", normal_style))
    content.append(Paragraph("• Membership is non-transferable and non-refundable.", normal_style))
    content.append(Spacer(1, 12))
    content.append(Paragraph("Thank you for choosing A1 Lift and Fit hub unisex Gym!", ParagraphStyle('ThankYou', parent=styles['Normal'], alignment=1, fontSize=12, spaceAfter=5)))
    content.append(Paragraph("Please keep this invoice for your records.", ParagraphStyle('Record', parent=styles['Normal'], alignment=1, fontSize=10)))

    # Build PDF once
    doc.build(content)
    pdf_buffer.seek(0)
    pdf_data = pdf_buffer.getvalue()

    if send_email:
        from django.core.mail import EmailMessage

        # Send email
        subject = f"Gym Membership Invoice - {member.name}"
        message = f"Dear {member.name},\n\nPlease find attached your gym membership invoice.\n\nThank you for choosing A1 Lift and Fit hub unisex Gym!\n\nBest regards,\nA1 Lift and Fit hub unisex Gym"
        email = EmailMessage(
            subject=subject,
            body=message,
            from_email=settings.DEFAULT_FROM_EMAIL,
            to=[member.email],
        )
        email.attach(f"invoice_{member.id}_{timezone.now().date()}.pdf", pdf_data, 'application/pdf')
        try:
            email.send()
            messages.success(request, f'Invoice PDF has been sent to {member.email}.')
        except Exception as e:
            messages.error(request, f'Failed to send email: {str(e)}')
        return redirect('member_list')
    else:
        # Return PDF for download
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="invoice_{member.id}_{timezone.now().date()}.pdf"'
        response.write(pdf_data)
        return response

